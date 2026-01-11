from openpyxl.styles import PatternFill, Alignment, Font

ADD_FILL = PatternFill("solid", fgColor="C6EFCE")
MODIFY_FILL = PatternFill("solid", fgColor="FFEB9C")
DELETE_FILL = PatternFill("solid", fgColor="FFC7CE")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)

WRAP_COLS = {"New_Value", "Reason", "Evidence_Text"}

def apply_table_formatting(ws, headers):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

def color_row(ws, row_idx: int, change_type: str):
    fill = None
    if change_type == "ADD":
        fill = ADD_FILL
    elif change_type == "MODIFY":
        fill = MODIFY_FILL
    elif change_type == "DELETE":
        fill = DELETE_FILL
    if fill:
        for c in range(1, ws.max_column + 1):
            ws.cell(row_idx, c).fill = fill

def normalize_sizes(ws, headers):
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 18
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = 18
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            name = headers[c-1]
            ws.cell(r, c).alignment = Alignment(
                wrap_text=(name in WRAP_COLS),
                vertical="top" if name in WRAP_COLS else "center"
            )
